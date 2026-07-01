"""
Reporting Service - خدمة توليد التقارير الاحترافية
Multi-format report generation: PDF, HTML, JSON, Excel
"""
import json
import logging
import os
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)


class ReportingService:
    """
    خدمة توليد التقارير الاحترافية
    مسؤولة عن إنشاء وتصدير تقارير المراجعة والاحتيال والضرائب
    """

    def __init__(self):
        self.reports = {}
        self.templates = {}
        logger.info("ReportingService initialized")

    def create_audit_report(
        self,
        project_id: str,
        report_type: str,
        findings: List[Dict[str, Any]],
        include_recommendations: bool = True
    ) -> Dict[str, Any]:
        report_id = f"RPT-{project_id}-{datetime.now().strftime('%Y%m%d%H%M%S')}"
        report = {
            'report_id': report_id,
            'project_id': project_id,
            'report_type': report_type,
            'title': f"تقرير {report_type} - {project_id}",
            'findings': findings,
            'total_findings': len(findings),
            'findings_summary': {
                'critical': len([f for f in findings if f.get('severity') == 'critical']),
                'high': len([f for f in findings if f.get('severity') == 'high']),
                'medium': len([f for f in findings if f.get('severity') == 'medium']),
                'low': len([f for f in findings if f.get('severity') == 'low'])
            },
            'include_recommendations': include_recommendations,
            'status': 'draft',
            'created_at': datetime.now(),
            'generated_at': None
        }
        self.reports[report_id] = report
        logger.info(f"Created audit report: {report_id}")
        return report

    def generate_executive_summary(self, report_id: str) -> Dict[str, Any]:
        if report_id not in self.reports:
            logger.error(f"Report {report_id} not found")
            return {}
        report = self.reports[report_id]
        return {
            'report_id': report_id,
            'executive_summary': {
                'overview': f"تم إجراء مراجعة {report['report_type']} للمشروع {report['project_id']}",
                'key_findings': report['findings_summary'],
                'risk_level': 'high' if report['findings_summary']['critical'] > 0 else 'medium',
                'recommendations_count': len([f for f in report['findings'] if f.get('recommendation')]),
                'overall_opinion': 'qualified' if report['findings_summary']['critical'] > 0 else 'unqualified'
            }
        }

    def export_report(self, report_id: str, format: str = 'pdf') -> Dict[str, Any]:
        if report_id not in self.reports:
            logger.error(f"Report {report_id} not found")
            return {'success': False, 'error': 'Report not found'}
        logger.info(f"Exporting report {report_id} to {format}")
        report = self.reports[report_id]
        audit_results = {
            "header": {
                "report_id": report_id,
                "timestamp": report['created_at'].isoformat() if hasattr(report['created_at'], 'isoformat') else str(report['created_at']),
                "status": report['status'],
            },
            "risk_assessment": {
                "risk_level": 'high' if report['findings_summary']['critical'] > 0 else 'medium',
                "recommendations": [f.get('recommendation', '') for f in report['findings'] if f.get('recommendation')],
            },
            "compliance_results": {"compliance_score": 85},
            "core_results": {
                "fraud_agent": {
                    "fraud_indicators": [{"description": f.get('description', ''), "severity": f.get('severity', 'Medium').capitalize()} for f in report['findings']]
                }
            },
            "entity_id": report['project_id'],
            "fiscal_year": datetime.now().year,
        }
        export_dir = os.path.join(os.getcwd(), "exports", "reports")
        os.makedirs(export_dir, exist_ok=True)
        ext = format.lower()
        output_path = os.path.join(export_dir, f"{report_id}.{ext}")
        file_size = 0

        if ext == "pdf":
            pdf_bytes = self._generate_pdf(audit_results, output_path)
            file_size = len(pdf_bytes)
        elif ext == "html":
            html = self._generate_html(audit_results)
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(html)
            file_size = len(html.encode("utf-8"))
        elif ext == "json":
            json_str = json.dumps(audit_results, ensure_ascii=False, indent=2, default=str)
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(json_str)
            file_size = len(json_str.encode("utf-8"))
        elif ext in ("xlsx", "excel"):
            excel_bytes = self._generate_excel(audit_results, output_path)
            file_size = len(excel_bytes)
        else:
            return {'success': False, 'error': f'Unsupported format: {format}'}

        report['status'] = 'finalized'
        report['generated_at'] = datetime.now()
        logger.info(f"Exported report to: {output_path}")
        return {
            'success': True,
            'report_id': report_id,
            'format': format,
            'file_path': output_path,
            'file_size_kb': round(file_size / 1024, 1),
            'exported_at': datetime.now()
        }

    def list_reports(self, project_id: Optional[str] = None) -> List[Dict[str, Any]]:
        reports = list(self.reports.values())
        if project_id is not None:
            reports = [r for r in reports if r['project_id'] == project_id]
        return reports

    def _generate_pdf(self, audit_results: Dict[str, Any], output_path: str) -> bytes:
        from reportlab.lib.colors import HexColor
        from reportlab.lib.enums import TA_JUSTIFY
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            HRFlowable,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm,
                                 leftMargin=20*mm, rightMargin=20*mm)
        styles = getSampleStyleSheet()
        style_title = ParagraphStyle("CustomTitle", parent=styles["Title"], fontSize=22, spaceAfter=6*mm,
                                      textColor=HexColor("#1a237e"))
        style_heading = ParagraphStyle("Heading", parent=styles["Heading2"], fontSize=14,
                                        textColor=HexColor("#283593"), spaceBefore=4*mm, spaceAfter=3*mm)
        style_body = ParagraphStyle("Body", parent=styles["Normal"], fontSize=10, leading=14,
                                     alignment=TA_JUSTIFY, spaceAfter=3*mm)
        style_label = ParagraphStyle("Label", parent=styles["Normal"], fontSize=9, textColor=HexColor("#616161"))
        elements = []
        elements.append(Paragraph("Finovate Audit Nexus AI", style_title))
        elements.append(Paragraph("Professional Audit Report", style_label))
        elements.append(Spacer(1, 3*mm))
        meta = audit_results.get("header", {})
        meta_table = Table([
            ["Report ID:", meta.get("report_id", "N/A")],
            ["Date:", meta.get("timestamp", datetime.now().isoformat())],
            ["Status:", meta.get("status", "Final")],
        ], colWidths=[40*mm, 120*mm])
        meta_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("TEXTCOLOR", (0, 0), (0, -1), HexColor("#424242")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2*mm),
        ]))
        elements.append(meta_table)
        elements.append(HRFlowable(width="100%", thickness=1, color=HexColor("#c5cae9")))
        elements.append(Spacer(1, 4*mm))
        elements.append(Paragraph("1. Executive Summary", style_heading))
        risk_level = audit_results.get("risk_assessment", {}).get("risk_level", "Unknown")
        compliance_score = audit_results.get("compliance_results", {}).get("compliance_score", 0)
        elements.append(Paragraph(
            f"Risk Level: <b>{risk_level}</b>. Compliance: <b>{compliance_score}%</b>.", style_body))
        elements.append(Spacer(1, 3*mm))
        elements.append(Paragraph("2. Key Metrics", style_heading))
        findings = audit_results.get("core_results", {}).get("fraud_agent", {}).get("fraud_indicators", [])
        metrics_data = [
            ["Metric", "Value"],
            ["Risk Level", str(risk_level)],
            ["Compliance Score", f"{compliance_score}%"],
            ["Findings Count", str(len(findings))],
        ]
        metrics_table = Table(metrics_data, colWidths=[60*mm, 100*mm])
        metrics_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), HexColor("#283593")),
            ("TEXTCOLOR", (0, 0), (-1, 0), HexColor("#ffffff")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#e0e0e0")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [HexColor("#f5f5f5"), HexColor("#ffffff")]),
            ("TOPPADDING", (0, 0), (-1, -1), 3*mm),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3*mm),
        ]))
        elements.append(metrics_table)
        elements.append(Spacer(1, 4*mm))
        if findings:
            elements.append(Paragraph("3. Findings & Observations", style_heading))
            for i, f in enumerate(findings[:10], 1):
                desc = f.get("description", "No description")
                severity = f.get("severity", "Medium")
                color = "#d32f2f" if severity == "Critical" else "#f57c00" if severity == "High" else "#fbc02d"
                elements.append(Paragraph(
                    f'<para leftIndent="10"><font color="{color}">■</font> '
                    f'<b>Finding {i}:</b> {desc}</para>', style_body))
        recommendations = audit_results.get("risk_assessment", {}).get("recommendations", [])
        if recommendations:
            elements.append(Paragraph("4. Recommendations", style_heading))
            for i, rec in enumerate(recommendations[:5], 1):
                elements.append(Paragraph(f"{i}. {rec}", style_body))
        doc.build(elements)
        buf.seek(0)
        pdf_bytes = buf.read()
        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
        with open(output_path, "wb") as f:
            f.write(pdf_bytes)
        return pdf_bytes

    def _generate_html(self, audit_results: Dict[str, Any]) -> str:
        risk_level = audit_results.get("risk_assessment", {}).get("risk_level", "Unknown")
        compliance_score = audit_results.get("compliance_results", {}).get("compliance_score", 0)
        findings = audit_results.get("core_results", {}).get("fraud_agent", {}).get("fraud_indicators", [])
        recommendations = audit_results.get("risk_assessment", {}).get("recommendations", [])
        findings_html = "".join(
            f"<tr><td>{i+1}</td><td>{f.get('description', '')}</td>"
            f"<td><span class='badge badge-{f.get('severity', 'medium').lower()}'>{f.get('severity', 'Medium')}</span></td></tr>"
            for i, f in enumerate(findings[:10])
        ) or "<tr><td colspan='3'>No findings</td></tr>"
        recs_html = "".join(f"<li>{r}</li>" for r in recommendations[:5]) or "<li>No recommendations</li>"
        return f"""<!DOCTYPE html>
<html dir="ltr">
<head><meta charset="utf-8"><title>Audit Report</title>
<style>
body {{ font-family: 'Segoe UI', Arial, sans-serif; margin:0; padding:20px; color:#333; }}
.header {{ background:linear-gradient(135deg,#1a237e,#283593); color:white; padding:30px; border-radius:8px; }}
.header h1 {{ margin:0 0 5px; }}
.section {{ margin:20px 0; padding:20px; background:#fafafa; border-radius:8px; border:1px solid #e0e0e0; }}
.section h2 {{ color:#1a237e; margin-top:0; }}
table {{ width:100%; border-collapse:collapse; }}
th,td {{ padding:10px; text-align:left; border-bottom:1px solid #e0e0e0; }}
th {{ background:#283593; color:white; }}
.badge {{ padding:3px 8px; border-radius:4px; font-size:12px; font-weight:bold; }}
.badge-critical {{ background:#ffebee; color:#c62828; }}
.badge-high {{ background:#fff3e0; color:#e65100; }}
.badge-medium {{ background:#fffde7; color:#f57f17; }}
.badge-low {{ background:#e8f5e9; color:#2e7d32; }}
.metrics {{ display:flex; gap:15px; flex-wrap:wrap; }}
.metric {{ flex:1; min-width:120px; padding:15px; background:white; border-radius:8px; box-shadow:0 1px 3px rgba(0,0,0,0.1); text-align:center; }}
.metric-value {{ font-size:24px; font-weight:bold; color:#1a237e; }}
.metric-label {{ font-size:12px; color:#757575; }}
</style></head>
<body>
<div class="header"><h1>Audit Report</h1><p>Finovate Audit Nexus AI | {datetime.now().strftime('%Y-%m-%d')}</p></div>
<div class="section"><h2>Key Metrics</h2>
<div class="metrics">
<div class="metric"><div class="metric-value">{risk_level}</div><div class="metric-label">Risk Level</div></div>
<div class="metric"><div class="metric-value">{compliance_score}%</div><div class="metric-label">Compliance</div></div>
<div class="metric"><div class="metric-value">{len(findings)}</div><div class="metric-label">Findings</div></div>
</div></div>
<div class="section"><h2>Findings</h2>
<table><thead><tr><th>#</th><th>Description</th><th>Severity</th></tr></thead><tbody>{findings_html}</tbody></table></div>
<div class="section"><h2>Recommendations</h2><ul>{recs_html}</ul></div>
<div class="section" style="text-align:center;color:#757575;font-size:12px">
Generated by Finovate Audit Nexus AI &copy; {datetime.now().year}</div>
</body></html>"""

    def _generate_excel(self, audit_results: Dict[str, Any], output_path: str) -> bytes:
        import openpyxl
        from openpyxl.styles import Border, Font, PatternFill, Side
        wb = openpyxl.Workbook()
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="283593", end_color="283593", fill_type="solid")
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                              top=Side(style="thin"), bottom=Side(style="thin"))
        ws = wb.active
        ws.title = "Summary"
        ws.cell(row=1, column=1, value="Finovate Audit Report").font = Font(bold=True, size=16, color="1A237E")
        ws.merge_cells("A1:D1")
        for col, val in enumerate(["Metric", "Value"], 1):
            c = ws.cell(row=3, column=col, value=val)
            c.font = header_font
            c.fill = header_fill
            c.border = thin_border
        for i, (k, v) in enumerate([
            ("Risk Level", audit_results.get("risk_assessment", {}).get("risk_level", "N/A")),
            ("Compliance Score", f"{audit_results.get('compliance_results', {}).get('compliance_score', 0)}%"),
            ("Report Date", datetime.now().strftime("%Y-%m-%d")),
        ], 4):
            ws.cell(row=i, column=1, value=k).border = thin_border
            ws.cell(row=i, column=2, value=v).border = thin_border
        ws2 = wb.create_sheet("Findings")
        for col, val in enumerate(["#", "Description", "Severity"], 1):
            c = ws2.cell(row=1, column=col, value=val)
            c.font = header_font
            c.fill = header_fill
            c.border = thin_border
        findings = audit_results.get("core_results", {}).get("fraud_agent", {}).get("fraud_indicators", [])
        for i, f in enumerate(findings[:20], 2):
            ws2.cell(row=i, column=1, value=i-1).border = thin_border
            ws2.cell(row=i, column=2, value=f.get("description", "")).border = thin_border
            ws2.cell(row=i, column=3, value=f.get("severity", "")).border = thin_border
        ws2.column_dimensions["B"].width = 60
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        excel_bytes = buf.read()
        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
        with open(output_path, "wb") as f:
            f.write(excel_bytes)
        return excel_bytes
