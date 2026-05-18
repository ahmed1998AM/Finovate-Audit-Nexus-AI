"""
Finovate Audit Nexus AI - Excel Connector
موصل Excel الاحترافي لاستيراد وتصدير البيانات المالية
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, List, Any, Optional, Union
from datetime import datetime
import os


class ExcelConnector:
    """
    موصل Excel الاحترافي
    يدعم قراءة وكتابة الملفات المالية المعقدة
    """
    
    def __init__(self):
        self.workbook = None
        self.dataframes = {}
        self.file_path = None
        
    def read_excel(self, file_path: str, sheet_names: Optional[List[str]] = None) -> Dict[str, pd.DataFrame]:
        """
        قراءة ملف Excel واستخراج البيانات
        
        Args:
            file_path: مسار الملف
            sheet_names: أسماء الأوراق المطلوبة (None = جميع الأوراق)
            
        Returns:
            قاموس يحتوي على بيانات كل ورقة كـ DataFrame
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
            
        self.file_path = file_path
        
        # قراءة جميع الأوراق
        if sheet_names is None:
            xl_file = pd.ExcelFile(file_path)
            sheet_names = xl_file.sheet_names
            
        results = {}
        for sheet in sheet_names:
            try:
                df = pd.read_excel(file_path, sheet_name=sheet)
                results[sheet] = df
                self.dataframes[sheet] = df
            except Exception as e:
                print(f"Warning: Could not read sheet '{sheet}': {e}")
                
        return results
    
    def read_trial_balance(self, file_path: str) -> pd.DataFrame:
        """قراءة ميزان المراجعة من Excel"""
        df = self.read_excel(file_path).values()
        if df:
            return list(df)[0]
        return pd.DataFrame()
    
    def read_journal_entries(self, file_path: str) -> pd.DataFrame:
        """قراءة قيود اليومية من Excel"""
        dfs = self.read_excel(file_path)
        # البحث عن ورقة تسمى Journal أو قيود
        for name, df in dfs.items():
            if 'journal' in name.lower() or 'قيود' in name.lower():
                return df
        # إرجاع أول ورقة إذا لم يتم العثور على ورقة محددة
        return list(dfs.values())[0] if dfs else pd.DataFrame()
    
    def write_financial_report(
        self,
        data: Dict[str, Any],
        output_path: str,
        report_title: str = "Financial Report"
    ) -> str:
        """
        كتابة تقرير مالي احترافي في Excel
        
        Args:
            data: البيانات المراد كتابتها
            output_path: مسار الملف الناتج
            report_title: عنوان التقرير
        """
        wb = Workbook()
        
        # تنسيقات احترافية
        title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        normal_font = Font(name='Arial', size=10)
        
        title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        alt_fill = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        center_alignment = Alignment(horizontal='center', vertical='center')
        right_alignment = Alignment(horizontal='right', vertical='center')
        
        # الورقة الأولى: الملخص
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # العنوان
        ws_summary.merge_cells('A1:E1')
        title_cell = ws_summary['A1']
        title_cell.value = report_title
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = center_alignment
        
        # تاريخ التقرير
        ws_summary['A2'] = f"Report Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws_summary['A2'].font = normal_font
        ws_summary.merge_cells('A2:E2')
        
        row_offset = 4
        
        # كتابة الأقسام المختلفة
        for section_name, section_data in data.items():
            if isinstance(section_data, dict):
                # إضافة عنوان القسم
                ws_summary.cell(row=row_offset, column=1).value = section_name.upper()
                ws_summary.cell(row=row_offset, column=1).font = header_font
                ws_summary.cell(row=row_offset, column=1).fill = header_fill
                
                row_offset += 1
                
                # كتابة البيانات
                for key, value in section_data.items():
                    ws_summary.cell(row=row_offset, column=1).value = key
                    ws_summary.cell(row=row_offset, column=2).value = value
                    
                    # تنسيق الخلايا
                    for col in range(1, 3):
                        cell = ws_summary.cell(row=row_offset, column=col)
                        cell.font = normal_font
                        cell.border = border
                        
                    row_offset += 1
                    
                row_offset += 2
                
        # ضبط عرض الأعمدة
        for col_idx in range(1, 6):
            column_letter = chr(64 + col_idx)  # A, B, C, D, E
            max_length = 0
            for row in ws_summary.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = min(max_length + 2, 50)
            ws_summary.column_dimensions[column_letter].width = adjusted_width
            
        # حفظ الملف
        wb.save(output_path)
        self.file_path = output_path
        
        return output_path
    
    def write_audit_findings(
        self,
        findings: List[Dict[str, Any]],
        output_path: str
    ) -> str:
        """
        كتابة نتائج التدقيق في Excel
        
        Args:
            findings: قائمة بنتائج التدقيق
            output_path: مسار الملف الناتج
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Audit Findings"
        
        # العناوين
        headers = ['ID', 'Type', 'Severity', 'Description', 'Value', 'Recommendation', 'Status']
        
        # تنسيق العناوين
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
        # كتابة البيانات
        for row_num, finding in enumerate(findings, 2):
            ws.cell(row=row_num, column=1, value=finding.get('id', row_num-1))
            ws.cell(row=row_num, column=2, value=finding.get('type', 'N/A'))
            ws.cell(row=row_num, column=3, value=finding.get('severity', 'N/A'))
            ws.cell(row=row_num, column=4, value=finding.get('description', 'N/A'))
            ws.cell(row=row_num, column=5, value=finding.get('value', 'N/A'))
            ws.cell(row=row_num, column=6, value=finding.get('recommendation', 'N/A'))
            ws.cell(row=row_num, column=7, value=finding.get('status', 'Open'))
            
            # تلوين حسب الخطورة
            severity = finding.get('severity', '').lower()
            if severity == 'high':
                fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            elif severity == 'medium':
                fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            else:
                fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                
            for col in range(1, 8):
                ws.cell(row=row_num, column=col).fill = fill
                
        # ضبط عرض الأعمدة
        column_widths = [10, 20, 12, 50, 15, 40, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
            
        wb.save(output_path)
        return output_path
    
    def export_dataframe(
        self,
        df: pd.DataFrame,
        output_path: str,
        sheet_name: str = "Data"
    ) -> str:
        """تصدير DataFrame إلى Excel"""
        df.to_excel(output_path, sheet_name=sheet_name, index=False)
        return output_path
    
    def validate_financial_data(self, df: pd.DataFrame) -> Dict[str, Any]:
        """
        التحقق من صحة البيانات المالية
        
        Returns:
            قاموس يحتوي على نتائج التحقق
        """
        validation_results = {
            'is_valid': True,
            'errors': [],
            'warnings': [],
            'row_count': len(df),
            'column_count': len(df.columns)
        }
        
        # التحقق من الأعمدة الأساسية
        required_columns = ['account', 'debit', 'credit']
        for col in required_columns:
            if col not in df.columns:
                validation_results['errors'].append(f"Missing required column: {col}")
                validation_results['is_valid'] = False
                
        # التحقق من القيم الفارغة
        null_counts = df.isnull().sum()
        for col, count in null_counts.items():
            if count > 0:
                validation_results['warnings'].append(
                    f"Column '{col}' has {count} null values"
                )
                
        # التحقق من أن القيم الرقمية صحيحة
        numeric_columns = ['debit', 'credit', 'balance']
        for col in numeric_columns:
            if col in df.columns:
                try:
                    pd.to_numeric(df[col], errors='raise')
                except:
                    validation_results['errors'].append(
                        f"Column '{col}' contains non-numeric values"
                    )
                    validation_results['is_valid'] = False
                    
        # التحقق من توازن القيود (إذا كان ممكنًا)
        if 'debit' in df.columns and 'credit' in df.columns:
            total_debit = pd.to_numeric(df['debit'], errors='coerce').sum()
            total_credit = pd.to_numeric(df['credit'], errors='coerce').sum()
            
            if abs(total_debit - total_credit) > 0.01:
                validation_results['warnings'].append(
                    f"Debit ({total_debit:.2f}) != Credit ({total_credit:.2f}). "
                    f"Difference: {abs(total_debit - total_credit):.2f}"
                )
                
        return validation_results


if __name__ == "__main__":
    # مثال اختباري
    print("=" * 60)
    print("Finovate Excel Connector - Test")
    print("=" * 60)
    
    connector = ExcelConnector()
    
    # إنشاء بيانات تجريبية
    test_data = {
        'summary': {
            'Total Assets': 1500000,
            'Total Liabilities': 600000,
            'Equity': 900000,
            'Revenue': 2000000,
            'Net Income': 250000
        },
        'ratios': {
            'Current Ratio': 1.67,
            'ROE': 27.78,
            'Debt to Equity': 66.67
        }
    }
    
    # كتابة تقرير
    output_file = "/workspace/exports/test_financial_report.xlsx"
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    
    connector.write_financial_report(
        data=test_data,
        output_path=output_file,
        report_title="Financial Analysis Report"
    )
    print(f"\n✅ Financial Report created: {output_file}")
    
    # إنشاء نتائج تدقيق تجريبية
    audit_findings = [
        {
            'id': 1,
            'type': 'Duplicate Entry',
            'severity': 'High',
            'description': 'تم اكتشاف قيد مكرر برقم 12345',
            'value': 50000,
            'recommendation': 'مراجعة وحذف القيد المكرر',
            'status': 'Open'
        },
        {
            'id': 2,
            'type': 'Missing Documentation',
            'severity': 'Medium',
            'description': 'فاتورة رقم 6789 بدون مستندات داعمة',
            'value': 15000,
            'recommendation': 'طلب المستندات الداعمة',
            'status': 'Pending'
        },
        {
            'id': 3,
            'type': 'Classification Error',
            'severity': 'Low',
            'description': 'خطأ في تصنيف المصروف',
            'value': 5000,
            'recommendation': 'إعادة التصنيف الصحيح',
            'status': 'Resolved'
        }
    ]
    
    findings_file = "/workspace/exports/test_audit_findings.xlsx"
    connector.write_audit_findings(
        findings=audit_findings,
        output_path=findings_file
    )
    print(f"✅ Audit Findings created: {findings_file}")
    
    # اختبار DataFrame
    df = pd.DataFrame({
        'Account': ['Cash', 'Accounts Receivable', 'Inventory', 'Fixed Assets'],
        'Debit': [100000, 200000, 150000, 500000],
        'Credit': [0, 0, 0, 0],
        'Balance': [100000, 200000, 150000, 500000]
    })
    
    df_file = "/workspace/exports/test_data.xlsx"
    connector.export_dataframe(df, df_file)
    print(f"✅ DataFrame exported: {df_file}")
    
    # التحقق من البيانات
    validation = connector.validate_financial_data(df)
    print(f"\n📋 Validation Results:")
    print(f"Valid: {validation['is_valid']}")
    print(f"Rows: {validation['row_count']}")
    print(f"Columns: {validation['column_count']}")
    if validation['warnings']:
        print(f"Warnings: {validation['warnings']}")
    if validation['errors']:
        print(f"Errors: {validation['errors']}")
    
    print("\n✅ Excel Connector Test Complete!")
